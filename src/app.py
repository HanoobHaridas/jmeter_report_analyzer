import streamlit as st
from tabulate import tabulate
import logging
import io
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
import os

from utils.file_handling import extract_zip
from utils.table_parsing import parse_jmeter_tables, REPORT_TABLES
from comparison.endpoint_comparison import compare_endpoint_stats
from comparison.aggregate_comparison import compare_aggregate_stats
from comparison.error_comparison import compare_errors

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Set page config
st.set_page_config(
    page_title="JMeter Report Analyzer",
    page_icon="📊",
    layout="wide"
)

def create_excel_with_multiple_sheets(tables):
    """Create an Excel file with multiple sheets, one for each table."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for table_type, df in tables.items():
            if df is not None and not df.empty:
                sheet_name = REPORT_TABLES[table_type][:31]  # Excel sheet names limited to 31 chars
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    return output.getvalue()

def create_markdown_report(tables):
    """Create a markdown report containing all tables."""
    markdown = []
    for table_type, df in tables.items():
        if df is not None and not df.empty:
            markdown.append(f"# {REPORT_TABLES[table_type]}\n")
            markdown.append(tabulate(df, headers='keys', tablefmt='pipe', showindex=False))
            markdown.append("\n\n")
    return "\n".join(markdown)

def create_comparison_excel(comparison_df, *report_names):
    """Create an Excel file for comparison data with dynamic colored metric groups for N reports."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Add a header sheet with comparison information
        header_df = pd.DataFrame({
            f'Report {i+1}': [name] for i, name in enumerate(report_names)
        })
        header_df.to_excel(writer, sheet_name='Comparison Info', index=False)
        comparison_df.to_excel(writer, sheet_name='Comparison', index=False)

        # Apply color formatting to metric groups
        workbook = writer.book
        worksheet = writer.sheets['Comparison']

        # Define a palette of subtle colors (extend as needed)
        palette = [
            "E3F0FD",  # Light blue
            "E2F7E1",  # Light green
            "FFFDE1",  # Light yellow
            "FDEEE3",  # Light peach
            "F3E1FD",  # Light purple
            "FDE1F0",  # Light pink
        ]

        columns = list(comparison_df.columns)
        n_reports = len(report_names)
        # The first column is 'Endpoint', so start from 1
        metric_group_indices = []
        i = 1
        while i < len(columns):
            metric_group_indices.append(list(range(i, min(i + n_reports, len(columns)))))
            i += n_reports

        # Apply alternating colors to each metric group
        for idx, group in enumerate(metric_group_indices):
            fill = PatternFill(start_color=palette[idx % len(palette)], end_color=palette[idx % len(palette)], fill_type="solid")
            for col in group:
                for row in range(2, worksheet.max_row + 1):  # +1 because openpyxl is 1-indexed and row 1 is header
                    worksheet.cell(row=row, column=col+1).fill = fill  # +1 because openpyxl columns are 1-indexed

    return output.getvalue()

def create_excel_with_multiple_comparison_sheets(comparison_dfs, report_names):
    """Create an Excel file with multiple comparison sheets for selected table types, with alternating blue/green coloring only."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Add a header sheet with comparison information
        header_df = pd.DataFrame({
            f'Report {i+1}': [name] for i, name in enumerate(report_names)
        })
        header_df.to_excel(writer, sheet_name='Comparison Info', index=False)
        # Write each comparison DataFrame to its own sheet
        color1 = "E3F0FD"  # Light blue
        color2 = "E2F7E1"  # Light green
        for table_type, df in comparison_dfs.items():
            if df is not None and not df.empty:
                sheet_name = table_type[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                columns = list(df.columns)
                n_reports = len(report_names)
                metric_group_indices = []
                i = 1
                while i < len(columns):
                    metric_group_indices.append(list(range(i, min(i + n_reports, len(columns)))))
                    i += n_reports
                for idx, group in enumerate(metric_group_indices):
                    fill = PatternFill(start_color=color1 if idx % 2 == 0 else color2, end_color=color1 if idx % 2 == 0 else color2, fill_type="solid")
                    for col in group:
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=col+1)
                            cell.fill = fill
    return output.getvalue()

def main():
    st.title("JMeter Report Analyzer and Comparer")
    
    # Mode selection
    mode = st.radio(
        "Select Mode",
        ["Single Report Analysis", "Report Comparison"]
    )
    
    if mode == "Single Report Analysis":
        st.header("Single Report Analysis")
        uploaded_file = st.file_uploader("Upload JMeter Report ZIP", type="zip")
        
        if uploaded_file is not None:
            if st.button("Analyze Report"):
                with st.spinner("Processing report..."):
                    try:
                        logger.debug("Starting report analysis...")
                        index_path, temp_dir = extract_zip(uploaded_file)
                        logger.debug(f"Extracted report to {index_path}")
                        
                        tables = parse_jmeter_tables(index_path)
                        logger.debug(f"Parsed tables: {tables}")
                        
                        # Check if we have any valid tables (not None and not empty)
                        has_valid_tables = any(df is not None and not df.empty for df in tables.values())
                        
                        if not has_valid_tables:
                            st.warning("No data found in the report. Please check if the report format is correct.")
                            return
                        
                        # Create download buttons for all formats
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            excel_data = create_excel_with_multiple_sheets(tables)
                            st.download_button(
                                label="📥 Download Excel Report",
                                data=excel_data,
                                file_name="jmeter_report.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        
                        with col2:
                            markdown_data = create_markdown_report(tables)
                            st.download_button(
                                label="📥 Download Markdown Report",
                                data=markdown_data,
                                file_name="jmeter_report.md",
                                mime="text/markdown"
                            )
                            
                    except Exception as e:
                        logger.error(f"Error processing report: {str(e)}", exc_info=True)
                        st.error(f"Error processing report: {str(e)}")
    
    else:  # Report Comparison
        st.header("Report Comparison")
        
        uploaded_reports = st.file_uploader("Upload JMeter Report ZIPs (multiple)", type="zip", accept_multiple_files=True)
        if 'report_names' not in st.session_state:
            st.session_state['report_names'] = []
        if 'names_confirmed' not in st.session_state:
            st.session_state['names_confirmed'] = False
        
        if uploaded_reports and len(uploaded_reports) > 1 and not st.session_state['names_confirmed']:
            with st.form("report_names_form"):
                temp_names = []
                for i, report in enumerate(uploaded_reports):
                    file_stem = os.path.splitext(report.name)[0]
                    name = st.text_input(f"Name for {report.name}", value=file_stem, key=f"report_name_{i}")
                    temp_names.append(name)
                submitted = st.form_submit_button("Confirm Report Names")
            if submitted:
                st.session_state['report_names'] = temp_names
                st.session_state['names_confirmed'] = True
                st.rerun()
            else:
                st.stop()
        
        if uploaded_reports and len(uploaded_reports) > 1 and st.session_state['names_confirmed']:
            report_names = st.session_state['report_names']
            st.markdown('**Selected Reports:**')
            for i, name in enumerate(report_names):
                st.markdown(f"- `{uploaded_reports[i].name}` → **{name}**")
            if st.button('Reset'):
                st.session_state['report_names'] = []
                st.session_state['names_confirmed'] = False
                st.rerun()
            table_types = st.multiselect(
                "Select Table Types for Comparison",
                list(REPORT_TABLES.values()),
                default=list(REPORT_TABLES.values())
            )
            if st.button("Compare Reports"):
                with st.spinner("Comparing reports..."):
                    try:
                        # Parse all reports once
                        all_tables = []
                        for report in uploaded_reports:
                            index_path, temp_dir = extract_zip(report)
                            tables = parse_jmeter_tables(index_path)
                            all_tables.append(tables)
                        # For each selected table type, build the comparison DataFrame
                        comparison_dfs = {}
                        for table_type in table_types:
                            table_key = [k for k, v in REPORT_TABLES.items() if v == table_type][0]
                            report_tables = [tables[table_key] if table_key in tables else None for tables in all_tables]
                            if any(df is None or df.empty for df in report_tables):
                                continue
                            if table_key == 'endpoint_stats':
                                comparison_df = compare_endpoint_stats(report_tables, report_names)
                            elif table_key == 'aggregate_summary':
                                comparison_df = compare_aggregate_stats(report_tables, report_names)
                            else:  # errors
                                comparison_df = compare_errors(report_tables[0], report_tables[1])
                            comparison_dfs[table_type] = comparison_df
                        if not comparison_dfs:
                            st.warning("No comparison data generated. Please check if the reports have matching data to compare.")
                            return
                        # Show all selected table types in the UI, aggregate first, errors last, skip empty errors
                        display_order = [REPORT_TABLES['aggregate_summary'], REPORT_TABLES['endpoint_stats'], REPORT_TABLES['errors']]
                        # Only show selected types, in the preferred order
                        shown_types = [t for t in display_order if t in table_types]
                        first_nonempty_type = None
                        for table_type in shown_types:
                            if table_type in comparison_dfs:
                                df = comparison_dfs[table_type]
                                if table_type == REPORT_TABLES['errors'] and (df is None or df.empty):
                                    continue
                                st.subheader(f"Comparison: {' vs '.join(report_names)} ({table_type})")
                                st.dataframe(df)
                                if first_nonempty_type is None:
                                    first_nonempty_type = table_type
                        # Create download buttons
                        col1, col2 = st.columns(2)
                        with col1:
                            excel_data = create_excel_with_multiple_comparison_sheets(comparison_dfs, report_names)
                            st.download_button(
                                label="📥 Download Excel Comparison",
                                data=excel_data,
                                file_name="comparison.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        with col2:
                            # Show markdown for the first non-empty type
                            if first_nonempty_type:
                                markdown_data = f"# Comparison: {' vs '.join(report_names)} ({first_nonempty_type})\n\n"
                                markdown_data += tabulate(comparison_dfs[first_nonempty_type], headers='keys', tablefmt='pipe', showindex=False)
                                st.download_button(
                                    label="📥 Download Markdown Comparison",
                                    data=markdown_data,
                                    file_name="comparison.md",
                                    mime="text/markdown"
                                )
                    except Exception as e:
                        logger.error(f"Error comparing reports: {str(e)}", exc_info=True)
                        st.error(f"Error comparing reports: {str(e)}")

if __name__ == "__main__":
    main() 